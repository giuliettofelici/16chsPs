from spi_2025A_display import *
from functools import partial
import sys, time, os
import RPi.GPIO as GPIO

# import form
from PyQt5.QtCore import (QThread,QTimer)
from PyQt5.QtWidgets import QMainWindow, QAction, QMenu, QApplication
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtGui import QPixmap


from PyQt5 import QtWidgets, uic
from PyQt5.QtWidgets import QMessageBox

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from _ast import While


my2025=SPI_2025A()

form_class = uic.loadUiType("/home/pi/Qbit/panel_2025a_display.ui")[0]    # Load the UI


list_span_txt=['0V to 5V','-2.5V to + 2.5V','-5V to +5V']
list_span_code=[0,1,2]
list_span_min=[0,-2.5,-5]
list_span_max=[5,2.5,5]
list_mon_mux=[0x10,0x11,0x12,0x13,0x14,0x15,0x16,0x17,0x18,0x19,0x1A,0x1B,0x1C,0x1D,0x1E,0x1F]
list_pB_wr_up=['pB_wr_up0', 'pB_wr_up1','pB_wr_up2', 'pB_wr_up3', 'pB_wr_up4','pB_wr_up5','pB_wr_up6','pB_wr_up7','pB_wr_up8', 'pB_wr_up9',
               'pB_wr_up10', 'pB_wr_up11', 'pB_wr_up12', 'pB_wr_up13','pB_wr_up14','pB_wr_up15']
list_pB_wr_up_t=['pB_wr_up0_t', 'pB_wr_up1_t','pB_wr_up2_t', 'pB_wr_up3_t', 'pB_wr_up4_t','pB_wr_up5_t','pB_wr_up6_t','pB_wr_up7_t','pB_wr_up8_t', 'pB_wr_up9_t',
               'pB_wr_up10_t', 'pB_wr_up11_t', 'pB_wr_up12_t', 'pB_wr_up13_t','pB_wr_up14_t','pB_wr_up15_t']
list_ledt_wr=['ledt_wr0', 'ledt_wr1','ledt_wr2', 'ledt_wr3', 'ledt_wr4','ledt_wr5','ledt_wr6','ledt_wr7','ledt_wr8', 'ledt_wr9',
                'ledt_wr10', 'ledt_wr11', 'ledt_wr12', 'ledt_wr13','ledt_wr14','ledt_wr15']
list_ledt_wr_t0=['ledt_wr0_t0', 'ledt_wr1_t0','ledt_wr2_t0', 'ledt_wr3_t0', 'ledt_wr4_t0','ledt_wr5_t0','ledt_wr6_t0','ledt_wr7_t0','ledt_wr8_t0', 'ledt_wr9_t0',
               'ledt_wr10_t0', 'ledt_wr11_t0', 'ledt_wr12_t0', 'ledt_wr13_t0','ledt_wr14_t0','ledt_wr15_t0']
list_ledt_wr_t1=['ledt_wr0_t1', 'ledt_wr1_t1','ledt_wr2_t1', 'ledt_wr3_t1', 'ledt_wr4_t1','ledt_wr5_t1','ledt_wr6_t1','ledt_wr7_t1','ledt_wr8_t1', 'ledt_wr9_t1',
               'ledt_wr10_t1', 'ledt_wr11_t1', 'ledt_wr12_t1', 'ledt_wr13_t1','ledt_wr14_t1','ledt_wr15_t1']
list_cBox_span=['cBox_span0', 'cBox_span1','cBox_span2','cBox_span3','cBox_span4','cBox_span5','cBox_span6','cBox_span7','cBox_span8',
                'cBox_span9','cBox_span10','cBox_span11', 'cBox_span12', 'cBox_span13', 'cBox_span14', 'cBox_span15']
list_cBox_span_t=['cBox_span0_t', 'cBox_span1_t','cBox_span2_t','cBox_span3_t','cBox_span4_t','cBox_span5_t','cBox_span6_t','cBox_span7_t','cBox_span8_t',
                'cBox_span9_t','cBox_span10_t','cBox_span11_t', 'cBox_span12_t', 'cBox_span13_t', 'cBox_span14_t', 'cBox_span15_t']
list_toggle_lB=['lB_ch0', 'lB_ch1', 'lB_ch2', 'lB_ch3', 'lB_ch4', 'lB_ch5', 'lB_ch6', 
                'lB_ch7', 'lB_ch8', 'lB_ch9', 'lB_ch10', 'lB_ch11', 'lB_ch12', 'lB_ch13', 'lB_ch14', 'lB_ch15']

list_pB_down=['pB_pd0', 'pB_pd1', 'pB_pd2', 'pB_pd3', 'pB_pd4', 'pB_pd5', 'pB_pd6', 'pB_pd7', 
              'pB_pd8', 'pB_pd9', 'pB_pd10', 'pB_pd11', 'pB_pd12', 'pB_pd13', 'pB_pd14', 'pB_pd15']

list_pB_down_t=['pB_pd0_t', 'pB_pd1_t', 'pB_pd2_t', 'pB_pd3_t', 'pB_pd4_t', 'pB_pd5_t', 'pB_pd6_t', 'pB_pd7_t', 
              'pB_pd8_t', 'pB_pd9_t', 'pB_pd10_t', 'pB_pd11_t', 'pB_pd12_t', 'pB_pd13_t', 'pB_pd14_t', 'pB_pd15_t']

list_cBox=['cBox_t0', 'cBox_t1', 'cBox_t2', 'cBox_t3', 'cBox_t4', 'cBox_t5', 'cBox_t6', 'cBox_t7', 'cBox_t8', 'cBox_t9', 
           'cBox_t10', 'cBox_t11', 'cBox_t12', 'cBox_t13', 'cBox_t14', 'cBox_t15']

list_toggle_sel = ["pB_toggle_s_0", "pB_toggle_s_1", "pB_toggle_s_2", "pB_toggle_s_3", "pB_toggle_s_4", "pB_toggle_s_5",
                   "pB_toggle_s_6", "pB_toggle_s_7", "pB_toggle_s_8", "pB_toggle_s_9", "pB_toggle_s_10", "pB_toggle_s_11",
                   "pB_toggle_s_12", "pB_toggle_s_13", "pB_toggle_s_14", "pB_toggle_s_15"]

# list_ledt_sense=['ledt_sen0','ledt_sen1','ledt_sen2','ledt_sen3','ledt_sen4','ledt_sen5','ledt_sen6',
#               'ledt_sen7', 'ledt_sen8','ledt_sen9','ledt_sen10','ledt_sen11','ledt_sen12','ledt_sen13',
#               'ledt_sen14','ledt_sen15']
# list_ledt_sense_t=['ledt_sen0_t','ledt_sen1_t','ledt_sen2_t','ledt_sen3_t','ledt_sen4_t','ledt_sen5_t','ledt_sen6_t',
#               'ledt_sen7_t', 'ledt_sen8_t','ledt_sen9_t','ledt_sen10_t','ledt_sen11_t','ledt_sen12_t','ledt_sen13_t',
#               'ledt_sen14_t','ledt_sen15_t']

GPIO.setwarnings(False)
GPIO.setmode(GPIO.BCM)  

print('------ start')

class MyWindowClass(QtWidgets.QMainWindow, form_class): # GUI Widjets 
    
    def __init__(self, parent=None):
        QtWidgets.QMainWindow.__init__(self, parent)
        self.setupUi(self)
        mywindowClass=MyWindowClass
        
        self.gpioSetBatt = 23
        GPIO.setup(self.gpioSetBatt, GPIO.OUT)
        
        self.gpioSenseBattCmd = 24
        GPIO.setup(self.gpioSenseBattCmd, GPIO.IN)
        
        #MENU
        self.actionSave_Setup.triggered.connect(self.save_setup)
        self.actionLoad_Setup.triggered.connect(self.load_setup)
      
        #self.actionSave_Setup.triggered.connect(lambda: self.clicked("New was clicked"))
        
        #BATTERY
        self.pB_battery_1.clicked.connect(self.battery)
        self.pB_battery_2.clicked.connect(self.battery)
        self.pB_battery_3.clicked.connect(self.battery)
        self.pB_battery_4.clicked.connect(self.battery)
        
        self.pB_battery_1.setStyleSheet("color: black; background-color: rgb(3,100,226);")
        self.pB_battery_1.setStyleSheet(
            "QPushButton#pB_battery_1:checked {color:black; background-color: rgb(0,191,255);}")
        self.pB_battery_2.setStyleSheet("color: black; background-color: rgb(3,100,226);")
        self.pB_battery_2.setStyleSheet(
            "QPushButton#pB_battery_2:checked {color:black; background-color: rgb(0,191,255);}")
        self.pB_battery_3.setStyleSheet("color: black; background-color: rgb(3,100,226);")
        self.pB_battery_3.setStyleSheet(
            "QPushButton#pB_battery_3:checked {color:black; background-color: rgb(0,191,255);}")
        self.pB_battery_4.setStyleSheet("color: black; background-color: rgb(3,100,226);")
        self.pB_battery_4.setStyleSheet(
            "QPushButton#pB_battery_4:checked {color:black; background-color: rgb(0,191,255);}")
        
        #POWER DOWN
        self.pB_pow_down_all_1.clicked.connect(self.set_pow_down_all)
        self.pB_pow_down_all_2.clicked.connect(self.set_pow_down_all)
        self.pB_pow_down_all_3.clicked.connect(self.set_pow_down_all)
        self.pB_pow_down_all_4.clicked.connect(self.set_pow_down_all)
 
        for i in range(len(list_pB_down)):
            exec('self.' + list_pB_down[i] +'.clicked.connect(partial(self.set_chan,' + str(i) + '))')
            exec('self.' + list_pB_down[i] + '.clicked.connect(self.power_down_n)')
            
        for i in range(len(list_pB_down_t)):
            exec('self.' + list_pB_down_t[i] +'.clicked.connect(partial(self.set_chan,' + str(i) + '))')
            exec('self.' + list_pB_down_t[i] + '.clicked.connect(self.power_down_n_t)')
            
        for i in range(len(list_toggle_sel)):
            exec('self.' + list_toggle_sel[i] +'.clicked.connect(partial(self.set_chan,' + str(i) + '))')
            exec('self.' + list_toggle_sel[i] + '.clicked.connect(self.toggle_selected)')
        
        #WRITE_UPDATE
        self.pB_write_all_1.clicked.connect(self.write_update_all)
        self.pB_write_all_2.clicked.connect(self.write_update_all)
        for i in range(len(list_pB_wr_up)):
            exec('self.' + list_pB_wr_up[i] +'.clicked.connect(partial(self.set_chan,' + str(i) + '))')
            exec('self.' + list_pB_wr_up[i] + '.clicked.connect(self.write_update_n)')
            exec('self.' + list_ledt_wr[i] +'.returnPressed.connect(partial(self.set_chan,' + str(i) + '))')
            exec('self.' + list_ledt_wr[i] + '.returnPressed.connect(self.write_update_n)')
        
        #WRITE TOGGLE
        self.pB_write_all_t1.clicked.connect(self.write_all_t)
        self.pB_write_all_t2.clicked.connect(self.write_all_t)
        for i in range(len(list_pB_wr_up_t)):
            exec('self.' + list_ledt_wr_t0[i] +'.returnPressed.connect(partial(self.set_chan,' + str(i) + '))')
            exec('self.' + list_ledt_wr_t0[i] + '.returnPressed.connect(self.write_single_toggle)')
            exec('self.' + list_ledt_wr_t1[i] +'.returnPressed.connect(partial(self.set_chan,' + str(i) + '))')
            exec('self.' + list_ledt_wr_t1[i] + '.returnPressed.connect(self.write_single_toggle)')
            exec('self.' + list_pB_wr_up_t[i] +'.clicked.connect(partial(self.set_chan,' + str(i) + '))')
            exec('self.' + list_pB_wr_up_t[i] + '.clicked.connect(self.write_single_toggle)')
         
        #MUX
        self.cBox_mon_mux_1.activated.connect(self.set_monitor_mux)
        self.cBox_mon_mux_2.activated.connect(self.set_monitor_mux)
        self.cBox_mon_mux_3.activated.connect(self.set_monitor_mux)
        self.cBox_mon_mux_4.activated.connect(self.set_monitor_mux)
        
        #SPAN
        self.cBox_Gspan_1.activated.connect(self.set_span_all)
        self.cBox_Gspan_2.activated.connect(self.set_span_all)
        
        #TOGGLE
        self.pB_toggle_1.setStyleSheet("color: black; background-color: rgb(3,100,226);")
        self.pB_toggle_1.setStyleSheet(
            "QPushButton#pB_toggle_1:checked {color:black; background-color: rgb(0,191,255);}")
        self.pB_toggle_2.setStyleSheet("color: black; background-color: rgb(3,100,226);")
        self.pB_toggle_2.setStyleSheet(
            "QPushButton#pB_toggle_2:checked {color:black; background-color: rgb(0,191,255);}")
    
        
        #TOGGLE SEL
        
        for i in range(len(list_toggle_sel)):
            exec('self.' + list_toggle_sel[i] + ".setStyleSheet(" + '"color:black;background-color:rgb(3,100,226);"' + ")")
            exec('self.' + list_toggle_sel[i] +  ".setStyleSheet(" + '"'  + "QPushButton#" + 
                 list_toggle_sel[i] + ":checked {color:black; background-color: rgb(0,191,255);}" + '"' + ")")
        
        
        self.pB_toggle_1.clicked.connect(self.toggle)
        self.pB_toggle_2.clicked.connect(self.toggle)
        
        
        #SPAN
        self.cBox_Gspan_1.activated.connect(self.set_span_all)
        self.cBox_Gspan_2.activated.connect(self.set_span_all)
        self.cBox_Gspan_3.activated.connect(self.set_span_all)
        self.cBox_Gspan_4.activated.connect(self.set_span_all)  
        for i in range(len(list_cBox_span)):
            exec('self.' + list_cBox_span[i] +'.activated.connect(partial(self.set_chan,' + str(i) + '))')
            exec('self.' + list_cBox_span[i] +'.activated.connect(partial(self.set_span_n))')
        
        self.init()
        
    def init(self):
        self.tabWidget.setCurrentIndex(0)    
        self.status_toggle=0
        self.index=0  
        self.file_data="datafile.xlsx"
        GPIO.output(self.gpioSetBatt, GPIO.LOW)
        self.pB_battery_1.setChecked(False)
        self.pB_battery_2.setChecked(False)
        self.pB_battery_3.setChecked(False)
        self.pB_battery_4.setChecked(False)
        if os.path.exists(self.file_data) == False:
            #print('GENERATE FILE')
            wb = Workbook()
            ws = wb.active
            ws.title = self.file_data
            m=1
            n=1
            for i in range(len(list_ledt_wr)):
                ws.cell(row=m, column=n+i, value=list_ledt_wr[i])
                ws.cell(row=m, column=n+i).alignment = Alignment(horizontal='center', vertical='center')
            m=3
            n=1
            for i in range(len(list_ledt_wr_t0)):
                ws.cell(row=m, column=n+i, value=list_ledt_wr_t0[i])
                ws.cell(row=m, column=n+i).alignment = Alignment(horizontal='center', vertical='center')
            m=5
            n=1
            for i in range(len(list_ledt_wr_t1)):
                ws.cell(row=m, column=n+i, value=list_ledt_wr_t1[i])
                ws.cell(row=m, column=n+i).alignment = Alignment(horizontal='center', vertical='center')
            m=7
            n=1
            for i in range(len(list_cBox_span)):
                ws.cell(row=m, column=n+i, value=list_cBox_span[i])
                ws.cell(row=m, column=n+i).alignment = Alignment(horizontal='center', vertical='center')
           
            wb.save(self.file_data)
            wb.close()   
       
        
            
            
    def alert(self,title, message):
        QtWidgets.QMessageBox.warning(self, title, message,
                    QtWidgets.QMessageBox.Close)
        
    def question(self, text_1, text_2):
        ret = QMessageBox.question(self, text_1, text_2, QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel, QMessageBox.Cancel)
        if ret == QMessageBox.Yes:
            return(True)
        else:
            return(False)
        
    def save_setup(self):
        reply=QtWidgets.QMessageBox.question(self, 'Save Setup','Save will overwrite previous values; do you want proceed ?' , 
                    QtWidgets.QMessageBox.Yes| QtWidgets.QMessageBox.No, QtWidgets.QMessageBox.No)
        if reply==QtWidgets.QMessageBox.Yes:
            wb=load_workbook(self.file_data)
            ws=wb.active
            row_ = 2
            for i in range(len(list_ledt_wr)):
                exec('global data_val; data_val=self.'+ list_ledt_wr[i]  + '.text()')
                ws.cell(column= i+1, row=row_, value=data_val)
            row_=4
            for i in range(len(list_ledt_wr_t0)):
                exec('global data_val_t; data_val_t=self.'+ list_ledt_wr_t0[i]  + '.text()')
                ws.cell(column= i+1, row=row_, value=data_val_t)  
            row_=6
            for i in range(len(list_ledt_wr_t1)):
                exec('global data_val_t; data_val_t=self.'+ list_ledt_wr_t1[i]  + '.text()')
                ws.cell(column= i+1, row=row_, value=data_val_t)  
            row_=8
            for i in range(len(list_cBox_span)):
                exec('global data_val_s; data_val_s=self.'+ list_cBox_span[i]  + '.currentIndex()')
                ws.cell(column= i+1, row=row_, value=data_val_s)  
            # row_=10
            # for i in range(len(list_cBox)):
            #     exec('global data_val_c; data_val_c=self.'+ list_cBox[i]  + '.isChecked()')
            #     ws.cell(column= i+1, row=row_, value=data_val_c)     
            wb.save(self.file_data)
            wb.close()
            
    
    def load_setup(self):
        reply=QtWidgets.QMessageBox.question(self, 'Load Setup','Save will overwrite current values; do you want proceed ?' , 
                    QtWidgets.QMessageBox.Yes| QtWidgets.QMessageBox.No, QtWidgets.QMessageBox.No)
        if reply==QtWidgets.QMessageBox.Yes:
            #print('----PROCEED----')
            ch_wr=[]
            ch_wr_t0=[]
            ch_wr_t1=[]
            ch_span=[]
            ch_box=[]
            wb=load_workbook(self.file_data)
            ws=wb.active
            for cell in ws[2]:
                ch_wr.append(cell.value)
            for cell in ws[4]:
                ch_wr_t0.append(cell.value)
            for cell in ws[6]:
                ch_wr_t1.append(cell.value)
            for cell in ws[8]:
                ch_span.append(cell.value)
            wb.close()
            for i in range(16):
                a=ch_span[i]
                exec('self.'+ list_ledt_wr[i]  + '.setText(str(ch_wr[i]))')
                exec('self.' + list_ledt_wr_t0[i] + '.setText(str(ch_wr_t0[i]))') 
                exec('self.' + list_ledt_wr_t1[i] + '.setText(str(ch_wr_t1[i]))')      
                exec('self.'+ list_cBox_span[i]  + '.setCurrentIndex(ch_span[i])')
                self.setup_apply(i)
    
    def setup_apply(self,ch):
        exec('global input_val; input_val=self.'+ list_ledt_wr[ch]  + '.text()')
        exec('global input_val1; input_val1=self.'+ list_ledt_wr_t0[ch]  + '.text()')
        exec('global input_val2; input_val2=self.'+ list_ledt_wr_t1[ch]  + '.text()')
        exec('global index; index=self.'+list_cBox_span[ch]+ '.currentIndex()')
        exec('global cbox; cbox=self.'+list_cBox[ch]+ '.isChecked()')
        span_min=list_span_min[self.index]
        span_max=list_span_max[self.index]
        span = list_span_code[self.index]
        num =float(input_val)
        num1 =float(input_val1)
        num2 =float(input_val2)
        if span_min==-2.5:
            val=num+2.5 
            val1 = num1+2.5
            val2 = num2+2.5
        elif span_min==-5:
            val=num+5
            val1 = num+5
            val2 = num+5       
        else:
            val=float(input_val)
            val1=float(input_val1)
            val2=float(input_val2)
        my2025.write_n(span,val,ch)
        self.power_down_status(False)
        if cbox:
            my2025.write_single_toggle_lib(span,val1,val2,ch)
        self.power_down_status(False)
   
    def toggle_selected(self,status):
        toggle_addr = 0
        
        toggle_addr =  2**self.chan
        
        print("TOGGLE", "canale=" , self.chan, "toggle_addr= ", toggle_addr)
               
        if status:
            my2025.toggle_sel(True, toggle_addr)
        else:
            my2025.toggle_sel(False, toggle_addr)

    
    def toggle(self, status):
        if status:
            my2025.toggle(True)
        else:
            my2025.toggle(False)
    
    
    def check_input(self,input_val, span_min, span_max):
        flag=True
        index=0
        if str(input_val):
            tmp=list(str(input_val))
            #print("FLAG, TMP", flag, tmp)
            for i in range(len(str(input_val))):
                if tmp[i] in "-0123456789.":
                    index=index+1
            if len(input_val) == index:
                val=float(input_val)
            else:
                flag=False
                self.alert("Alert", "Input value must be a number")
                return(flag)
        else:    
            self.alert("Alert", "yuo must insert a value")
            flag=False
            return(flag)
        if flag:
            if val>=span_min and val <= span_max:
                return(True)
            else:
                self.alert("Alert","Value exceeds input range")
                return(False)  
            
    def check_input_t(self,input_val1, input_val2, span_min, span_max):
        flag1=True
        flag1_and=False
        flag2=True
        flag2_and=False
        index1=0
        index2=0
        if str(input_val1):
            tmp=list(str(input_val1))
            for i in range(len(str(input_val1))):
                if tmp[i] in "-0123456789.":
                    index1=index1+1
            if len(input_val1) == index1:
                val1=float(input_val1)
            else:
                flag1=False
                self.alert("Alert", "Value input must be a number")
        else:    
            self.alert("Alert", "You must insert a value in Value field")
            flag1=False
        
        if flag1:
            if val1<=span_min or val1 >= span_max:
                self.alert("Alert","Value exceeds input range")
                flag1_and=False
            else:
                flag1_and=True
                #print("TRUE_")
           
            
        if str(input_val2):
            tmp=list(str(input_val2))
            for i in range(len(str(input_val2))):
                if tmp[i] in "-0123456789.":
                    index2=index2+1
            if len(input_val2) == index2:
                val2=float(input_val2)
            else:
                flag2=False
                self.alert("Alert", "Toggle input must be a number")
        else:    
            self.alert("Alert", "You must insert a value in Toggle field")
            flag2=False
           
        if flag2:
            if val2<=span_min or val2 >= span_max:
                self.alert("Alert","Toggle value exceeds input range")
                flag2_and=False
            else:
                flag2_and=True    
        return(flag1_and, flag2_and)
                                         
                    
    def set_chan(self,selectchan): # BIND to the selected field  
        self.chan = selectchan   # Identify the return pressed
    
    def set_span_all(self):
        tab=self.tabWidget.currentIndex()
        if tab == 0:
            self.index=self.cBox_Gspan_1.currentIndex()
            self.cBox_Gspan_2.setCurrentIndex(self.index)
            self.cBox_Gspan_3.setCurrentIndex(self.index)
            self.cBox_Gspan_4.setCurrentIndex(self.index)
        if tab == 1:
            self.index=self.cBox_Gspan_2.currentIndex()
            self.cBox_Gspan_1.setCurrentIndex(self.index)
            self.cBox_Gspan_3.setCurrentIndex(self.index)
            self.cBox_Gspan_4.setCurrentIndex(self.index)
        if tab == 2:
            self.index=self.cBox_Gspan_3.currentIndex()
            self.cBox_Gspan_1.setCurrentIndex(self.index)
            self.cBox_Gspan_2.setCurrentIndex(self.index)
            self.cBox_Gspan_4.setCurrentIndex(self.index)
        if tab == 3:
            self.index=self.cBox_Gspan_4.currentIndex()
            self.cBox_Gspan_1.setCurrentIndex(self.index)
            self.cBox_Gspan_2.setCurrentIndex(self.index)
            self.cBox_Gspan_3.setCurrentIndex(self.index)
        for i in range(16):
            exec('self.'+list_cBox_span[i]+'.setCurrentIndex('+ str(self.index) +')')
            exec('self.'+list_cBox_span_t[i]+'.setCurrentIndex('+ str(self.index) +')')
        self.gspan = list_span_code[self.index]
        my2025.set_span_all(self.gspan)  
        
    def write_all_ledt(self, val): 
        self.ledt_write_all_1.setText(val)
        self.ledt_write_all_2.setText(val)
        self.ledt_write_all_3.setText(val)
        self.ledt_write_all_7.setText(val)
     
         
    def set_span_n(self):
        addr=self.chan
        exec('global code; code=self.'+ list_cBox_span[self.chan]  + '.currentIndex()')
        my2025.set_span_n(code, addr)
        
    def power_down_n(self):
        if self.question("Warning",  "Power Down Chip will clear channel settings, do you want proceed ?"):
            addr=self.chan 
            my2025.power_down_n(addr)
            exec("self." + list_ledt_wr[addr] + ".setText('0')")
        
    def power_down_n_t(self):
        if self.question("Warning",  "Power Down Chip will clear all settings, do you want proceed ?"):
            addr=self.chan 
            my2025.power_down_n(addr)
            exec("self." + list_ledt_wr[addr] + ".setText('0')")
            exec("self." + list_ledt_wr_t0[addr] + ".setText('0')")
            exec("self." + list_ledt_wr_t1[addr] + ".setText('0')")
        
                
    def span_min_max(self):
        span_min=list_span_min[self.index]
        span_max=list_span_max[self.index]
        return(span_min, span_max)
    
    def set_monitor_mux(self):
        
        print(" ----------- Enter Monitor Mux")

        tab=self.tabWidget.currentIndex()
        if tab==0:
            print(" ----------- Enter Monitor Mux 1")
            index=self.cBox_mon_mux_1.currentIndex()
            my2025.monitor_mux(list_mon_mux[index-1], index)
        elif tab == 1:
            index=self.cBox_mon_mux_2.currentIndex()
            my2025.monitor_mux(list_mon_mux[index-1], index)
        elif tab==3:
            index=self.cBox_mon_mux_3.currentIndex()
            my2025.monitor_mux(list_mon_mux[index-1], index)
        else:
            index=self.cBox_mon_mux_4.currentIndex()
            my2025.monitor_mux(list_mon_mux[index-1], index)
        
    def set_pow_down_all(self,status):
        if self.question("Warning",  "Power Down Chip will clear all settings, do you want proceed ?"):
            my2025.power_down_all()
            self.ledt_write_all_1.setText('0')
            self.ledt_write_all_2.setText('0')
            self.ledt_write_all_3.setText('0')
            self.ledt_write_all_7.setText('0')
            self.ledt_write_all_t1.setText('0')
            self.ledt_write_all_t2.setText('0')
            for i in range(len(list_ledt_wr)):
                exec('self.'+list_ledt_wr[i]+'.setText("0")')
                exec('self.'+list_ledt_wr_t0[i]+'.setText("0")')
                exec('self.'+list_ledt_wr_t1[i]+'.setText("0")')
            
    def write_update_all(self):
        tab=self.tabWidget.currentIndex()
        if tab==0:
            self.index=self.cBox_Gspan_1.currentIndex()
            self.pB_toggle_1.setChecked(False)
            self.pB_toggle_2.setChecked(False)
            input_val=self.ledt_write_all_1.text()
        elif tab==1:
            self.index=self.cBox_Gspan_2.currentIndex()
            self.pB_toggle_1.setChecked(False)
            self.pB_toggle_2.setChecked(False)
            input_val=self.ledt_write_all_2.text()
        self.set_span_all()
        self.write_all_ledt(input_val)
        self.gspan = list_span_code[self.index]
        span_min,span_max=self.span_min_max()     
        if self.check_input(input_val,span_min, span_max):
            num =float(input_val)
            if span_min==-2.5: 
                val = num+2.5
            elif span_min==-5:
                val = num+5             
            else:
                val=float(input_val)
            my2025.write_all(self.gspan,val)  
            #self.power_down_status(False)
            for i in range(len(list_ledt_wr)):
                exec('self.'+list_ledt_wr[i]+'.setText(str(num))')
                exec('self.'+list_ledt_wr_t0[i]+'.setText(str(num))')
        else:
            self.write_all_ledt('         ')
           
    def write_all_t(self):
        tab=self.tabWidget.currentIndex()
        if tab==2:
            input_val1=self.ledt_write_all_3.text()
            input_val2=self.ledt_write_all_t1.text()
            self.ledt_write_all_t2.setText(input_val1)
        elif tab==3:
            input_val1=self.ledt_write_all_7.text()
            input_val2=self.ledt_write_all_t2.text()
            self.ledt_write_all_t1.setText(input_val2)
        self.set_span_all()
        self.write_all_ledt(input_val1)
        self.gspan = list_span_code[self.index]
        span_min,span_max=self.span_min_max()
        print("CHECK INPUT", "val1= ", input_val1, "val2= ", input_val2, "span_min= ", span_min, "span_max= ", span_max)
        flag1_and, flag2_and =self.check_input_t(input_val1,input_val2,span_min, span_max)
        if (flag1_and and flag2_and):
            num1 =float(input_val1)
            num2= float(input_val2)
            if span_min==-2.5: 
                val1 = num1+2.5
                val2=num2+2.5
            elif span_min==-5:
                val1 = num1+5
                val2=num2+5             
            else:
                val1=float(input_val1)
                val2=float(input_val2)
            my2025.write_all_toggle(self.gspan,val1, val2)
            for i in range(len(list_ledt_wr)):
                exec('self.'+list_ledt_wr[i]+'.setText(str(num1))')
                exec('self.'+list_ledt_wr_t0[i]+'.setText(str(num1))')
                exec('self.'+list_ledt_wr_t1[i]+'.setText(str(num2))')
        else:
            self.ledt_write_all_1.setText('')
            self.ledt_write_all_2.setText('')
            self.ledt_write_all_3.setText('')
            self.ledt_write_all_7.setText('')
            self.ledt_write_all_t1.setText('')
            self.ledt_write_all_t2.setText('')
    
    
    def battery(self, status):
        if status:
            GPIO.output(self.gpioSetBatt, GPIO.HIGH)
            self.pB_battery_1.setChecked(True)
            self.pB_battery_2.setChecked(True)
            self.pB_battery_3.setChecked(True)
            self.pB_battery_4.setChecked(True)
            #print('>>>> BATT >>>> HIGH')
        else:
            GPIO.output(self.gpioSetBatt, GPIO.LOW)
            self.pB_battery_1.setChecked(False)
            self.pB_battery_2.setChecked(False)
            self.pB_battery_3.setChecked(False)
            self.pB_battery_4.setChecked(False)
            #print('>>>> BATT >>>> LOW')
    
            
    def update_all(self):
        my2025.update_all()
            
    def write_update_n(self):
        tab=self.tabWidget.currentIndex()
        if tab==0:
            self.index=self.cBox_Gspan_1.currentIndex()
        elif tab==1:
            self.index=self.cBox_Gspan_2.currentIndex()
        exec('global input_val; input_val=self.'+ list_ledt_wr[self.chan]  + '.text()')
        exec('global index; index=self.'+list_cBox_span[self.chan]+ '.currentIndex()')
        tag=list_ledt_wr[self.chan]
        addr= self.chan
        span_min=list_span_min[index]
        span_max=list_span_max[index]
        span = list_span_code[index]
        if self.check_input(input_val,span_min, span_max):
            num =float(input_val)
            if span_min==-2.5: 
                val = num+2.5
            elif span_min==-5:
                val = num+5      
            else:
                val=float(input_val)
            my2025.write_n(span,val,addr)
            exec('global input_val; self.' + list_ledt_wr_t0[self.chan] + '.setText(input_val)')
           
        else:
            exec('self.' + list_ledt_wr[self.chan] + '.setText("")')    
    
    
    def write_single_toggle(self):
        
        print("")
        print(">>>>>> WRITE SINGLE TOGGLE <<<<<<")
        
        
        exec('global input_val1; input_val1=self.'+ list_ledt_wr_t0[self.chan]  + '.text()')
        exec('global input_val2; input_val2=self.'+ list_ledt_wr_t1[self.chan]  + '.text()')
        exec('global index; index=self.'+list_cBox_span_t[self.chan]+ '.currentIndex()')
        toggle_addr=self.chan
        span_min=list_span_min[index]
        span_max=list_span_max[index]
        span = list_span_code[index]
        flag1_and, flag2_and = self.check_input_t(input_val1,input_val2, span_min, span_max)
        if (flag1_and and flag2_and):
            num1 =float(input_val1)
            num2 = float(input_val2)
            if span_min==-2.5: 
                val1 = num1+2.5
                val2=num2+2.5
            elif span_min==-5:
                    val1 = num1+5 
                    val2= num2+5     
            else:
                val1=float(input_val1)
                val2=float(input_val2)
                
                print("VAL1, VAL2", val1, val2)        
            my2025.write_single_toggle_lib(span,val1,val2,self.chan, toggle_addr)
        else:
            exec('self.' + list_ledt_wr_t1[self.chan] + '.setText("")') 
            exec('self.' + list_ledt_wr_t0[self.chan] + '.setText("")')
       
            
            
    def convert_single(self,addr,num):
        tab=self.tabWidget.currentIndex()
        my2025.monitor_mux(tab, list_mon_mux[addr],addr) 
        if num<0:
            data_conv=-1*float(my2025.read_ADC(1))
        else:
            data_conv=float(my2025.read_ADC(0))    
        data_conv=float("{:.3f}".format(data_conv))
        if tab==0:
            exec('self.'+list_ledt_sense[addr]+'.setText(str(data_conv))')
        else:
            exec('self.'+list_ledt_sense_t[addr]+'.setText(str(data_conv))')
 
            
    def convert_all(self):
        tab=self.tabWidget.currentIndex()
        if tab==0:
            temp=self.ledt_write_all.text()
            if temp=='':
                self.alert('Missing Data', 'Global set data missing')
                return
            else:
                val_set=float(temp)
        else:
            temp=self.ledt_write_all_t.text()
            if temp=='':
                self.alert('Missing Data', 'Global set data missing')
                return
            else:
                val_set=float(temp)
        index_span=self.cBox_Gspan.currentIndex()
        for i in range(16):
            my2025.monitor_mux(tab, list_mon_mux[i],i)
            if val_set<0:
                data_conv=-1*float(my2025.read_ADC(1))
                data_conv=float("{:.3f}".format(data_conv))
            else:
                data_conv=float(my2025.read_ADC(0))
                data_conv=float("{:.3f}".format(data_conv))
        
            if tab==0:                 
                exec('self.'+list_ledt_sense[i]+'.setText(str(data_conv))')
            else:
                exec('self.'+list_ledt_sense_t[i]+'.setText(str(data_conv))')
        
app = QtWidgets.QApplication(sys.argv)
myWindow = MyWindowClass()
myWindow.show()
app.exec_()

    
